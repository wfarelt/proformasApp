# Descargar plantilla de inventario inicial
from openpyxl import Workbook

def download_inventory_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "inventario"
    sheet.append(["product_code", "quantity", "cost", "precio", "location", "descripcion"])
    sheet.append(["1R0750", 1, 10, 16, "a-1", "Filtro de aceite"])
    sheet.append(["2R0800", 2, 20, 30, "b-2"])
    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_inventario_inicial.xlsx"'
    return response
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.contrib import messages
from django.utils.timezone import now
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from datetime import timedelta, time
from django.db.models import Q, F, OuterRef, Subquery, Value, IntegerField
from django.db.models.functions import Coalesce
from datetime import datetime
from django.utils.timezone import make_aware
from django.utils import timezone
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse

from core.models import Detalle, Proforma, ProductStock, Warehouse
from .models import Producto, Purchase, PurchaseDetail, Movement, MovementItem
from .forms import PurchaseForm, PurchaseDetailFormSet, MovementForm, MovementItemFormSet, InventoryUploadForm, WarehouseForm

from django.db import transaction
from django.contrib.contenttypes.models import ContentType

# InventoryUploadForm
import openpyxl
from django.db.models import Sum, Count

from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from io import BytesIO

from inv.services.purchase_confirmation_service import confirm_purchase_and_apply_inventory
from inv.services.warehouse_transfer_service import create_warehouse_transfer
from core.services.inventory_service import InsufficientWarehouseStock, apply_warehouse_stock_change
from core.services.warehouse_access_service import WarehouseAccessDenied, accessible_warehouses, default_user_warehouse, resolve_user_warehouse
import json


# INGRESOS


def _can_manage_warehouses(user):
    return getattr(user, 'can_manage_all_warehouses', False)


def _selected_warehouse_for_user(request):
    warehouse_id = request.GET.get('warehouse_id')
    if request.user.can_manage_all_warehouses and not warehouse_id:
        return None
    return resolve_user_warehouse(request.user, warehouse_id)


def _accessible_purchases(user):
    purchases = Purchase.objects.all()
    if not user.can_manage_all_warehouses:
        purchases = purchases.filter(warehouse=default_user_warehouse(user))
    return purchases


def _accessible_movements(user):
    movements = Movement.objects.all()
    if not user.can_manage_all_warehouses:
        movements = movements.filter(warehouse=default_user_warehouse(user))
    return movements


@login_required(login_url='login')
def warehouse_list(request):
    if not request.user.can_access_inventory:
        messages.error(request, 'No tienes permisos para consultar almacenes.')
        return redirect('home')

    warehouses = (
        accessible_warehouses(request.user)
        .annotate(
            product_count=Count('product_stocks', distinct=True),
            total_quantity=Sum('product_stocks__quantity'),
        )
        .order_by('-is_default', 'name')
    )
    return render(request, 'inv/warehouse/warehouse_list.html', {
        'title': 'Almacenes',
        'warehouses': warehouses,
    })


@login_required(login_url='login')
def warehouse_create(request):
    if not _can_manage_warehouses(request.user):
        messages.error(request, 'No tienes permisos para administrar almacenes.')
        return redirect('home')

    form = WarehouseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        warehouse = form.save(commit=False)
        if Warehouse.objects.filter(code=warehouse.code).exists():
            form.add_error('code', 'Ya existe un almacén con este código.')
        else:
            with transaction.atomic():
                if warehouse.is_default:
                    Warehouse.objects.update(is_default=False)
                warehouse.save()
            messages.success(request, 'Almacén creado correctamente.')
            return redirect('warehouse_list')

    return render(request, 'inv/warehouse/warehouse_form.html', {
        'title': 'Nuevo almacén',
        'form': form,
    })


@login_required(login_url='login')
def warehouse_edit(request, pk):
    if not _can_manage_warehouses(request.user):
        messages.error(request, 'No tienes permisos para administrar almacenes.')
        return redirect('home')

    warehouse = get_object_or_404(Warehouse, pk=pk)

    form = WarehouseForm(request.POST or None, instance=warehouse)
    if request.method == 'POST' and form.is_valid():
        with transaction.atomic():
            warehouse = form.save(commit=False)
            if warehouse.is_default:
                Warehouse.objects.exclude(pk=warehouse.pk).update(is_default=False)
            warehouse.save()
        messages.success(request, 'Almacén actualizado correctamente.')
        return redirect('warehouse_list')

    return render(request, 'inv/warehouse/warehouse_form.html', {
        'title': 'Editar almacén',
        'form': form,
        'warehouse': warehouse,
    })


@login_required(login_url='login')
def warehouse_detail(request, pk):
    if not request.user.can_access_inventory:
        messages.error(request, 'No tienes permisos para consultar almacenes.')
        return redirect('home')

    warehouse = get_object_or_404(accessible_warehouses(request.user), pk=pk)
    query = request.GET.get('q', '').strip()
    stocks = ProductStock.objects.filter(warehouse=warehouse).select_related('product', 'product__brand')
    if query:
        stocks = stocks.filter(
            Q(product__nombre__icontains=query)
            | Q(product__referencia_cruzada__icontains=query)
            | Q(product__descripcion__icontains=query)
        )
    stocks = stocks.order_by('product__nombre')
    paginator = Paginator(stocks, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    totals = stocks.aggregate(total_quantity=Sum('quantity'), product_count=Count('product', distinct=True))
    return render(request, 'inv/warehouse/warehouse_detail.html', {
        'title': warehouse.name,
        'warehouse': warehouse,
        'page_obj': page_obj,
        'query': query,
        'total_quantity': totals['total_quantity'] or 0,
        'product_count': totals['product_count'] or 0,
    })

@login_required       
def product_search(request):
    query = request.GET.get('q', '')
    products = Producto.objects.filter(
        Q(nombre__icontains=query) | Q(referencia_cruzada__icontains=query)
    )[:10]  # Muestra solo 10 resultados
    data = [{"id": p.id, "name": p.nombre} for p in products]
    return JsonResponse(data, safe=False)

# REPORTES

@login_required
def reporte_analitica_productos(request):
    dias_permitidos = [7, 15, 30, 60]
    tipo_permitidos = {
        'mas_vendidos': 'Productos más vendidos',
        'rotacion_sin_stock': 'Productos con rotación sin stock',
        'menos_vendidos_con_stock': 'Productos menos vendidos con stock',
        'otros': 'Productos sin movimiento con stock',
    }

    try:
        dias = int(request.GET.get('dias', 15))
    except (TypeError, ValueError):
        dias = 15

    if dias not in dias_permitidos:
        dias = 15

    tipo = request.GET.get('tipo', 'mas_vendidos')
    if tipo not in tipo_permitidos:
        tipo = 'mas_vendidos'

    fecha_limite = now() - timedelta(days=dias)
    warehouses = accessible_warehouses(request.user)
    selected_warehouse = _selected_warehouse_for_user(request)
    warehouse_stock = ProductStock.objects.filter(
        product_id=OuterRef('producto_id'),
        warehouse=selected_warehouse,
    )
    stock_actual = Coalesce(
        Subquery(warehouse_stock.values('quantity')[:1]),
        Value(0),
        output_field=IntegerField(),
    )
    ubicacion_actual = Coalesce(Subquery(warehouse_stock.values('location')[:1]), Value(''))

    detalles_base = Detalle.objects.filter(
        proforma__estado='EJECUTADO',
        proforma__fecha__gte=fecha_limite,
    )

    if getattr(request.user, 'company_id', None):
        detalles_base = detalles_base.filter(proforma__company=request.user.company)
    if selected_warehouse:
        detalles_base = detalles_base.filter(proforma__warehouse=selected_warehouse)

    if tipo == 'mas_vendidos':
        resultados = (
            detalles_base
            .values('producto_id')
            .annotate(
                codigo=F('producto__nombre'),
                descripcion=F('producto__descripcion'),
                stock_actual=stock_actual if selected_warehouse else F('producto__stock'),
                ubicacion=ubicacion_actual if selected_warehouse else F('producto__location'),
                indicador=Sum('cantidad'),
            )
            .order_by('-indicador', 'codigo')
        )
    elif tipo == 'rotacion_sin_stock':
        resultados = (
            detalles_base
            .values('producto_id')
            .annotate(
                codigo=F('producto__nombre'),
                descripcion=F('producto__descripcion'),
                stock_actual=stock_actual if selected_warehouse else F('producto__stock'),
                ubicacion=ubicacion_actual if selected_warehouse else F('producto__location'),
                indicador=Sum('cantidad'),
            )
        )
        resultados = resultados.filter(stock_actual__lte=0).order_by('-indicador', 'codigo')
    elif tipo == 'menos_vendidos_con_stock':
        resultados = (
            detalles_base
            .values('producto_id')
            .annotate(
                codigo=F('producto__nombre'),
                descripcion=F('producto__descripcion'),
                stock_actual=stock_actual if selected_warehouse else F('producto__stock'),
                ubicacion=ubicacion_actual if selected_warehouse else F('producto__location'),
                indicador=Sum('cantidad'),
            )
        )
        resultados = resultados.filter(stock_actual__gt=0).order_by('indicador', 'codigo')
    else:
        productos_con_movimiento = detalles_base.values_list('producto_id', flat=True).distinct()
        if selected_warehouse:
            productos_base = Producto.objects.filter(
                warehouse_stocks__warehouse=selected_warehouse,
                warehouse_stocks__quantity__gt=0,
            )
            stock_annotation = F('warehouse_stocks__quantity')
            location_annotation = F('warehouse_stocks__location')
        else:
            productos_base = Producto.objects.filter(stock__gt=0)
            stock_annotation = F('stock')
            location_annotation = F('location')
        resultados = (
            productos_base
            .exclude(id__in=productos_con_movimiento)
            .annotate(
                producto_id=F('id'),
                codigo=F('nombre'),
                stock_actual=stock_annotation,
                ubicacion=location_annotation,
                indicador=Value(0, output_field=IntegerField()),
            )
            .values('producto_id', 'codigo', 'descripcion', 'stock_actual', 'ubicacion', 'indicador')
            .order_by('codigo')
        )

    resultados = resultados[:100]
    paginator = Paginator(resultados, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'inv/reports/analitica_productos.html', {
        'title': 'Analitica de productos',
        'page_obj': page_obj,
        'dias': dias,
        'tipo': tipo,
        'dias_permitidos': dias_permitidos,
        'tipo_permitidos': tipo_permitidos,
        'tipo_label': tipo_permitidos[tipo],
        'indicador_label': 'Cantidad vendida' if tipo != 'otros' else 'Movimientos en el periodo',
        'warehouses': warehouses,
        'selected_warehouse': selected_warehouse,
    })

def historial_ventas_producto(request):
    producto_id = request.GET.get("producto_id")
    dias = int(request.GET.get("dias", 30))  # Rango de días (por defecto, 30 días)
    warehouses = accessible_warehouses(request.user)
    selected_warehouse = _selected_warehouse_for_user(request)

    productos = Producto.objects.all()  # Para llenar el select de productos
    ventas = []

    if producto_id:
        fecha_limite = now() - timedelta(days=dias)
        ventas = (
            Detalle.objects
            .filter(producto_id=producto_id, proforma__estado="EJECUTADO", proforma__fecha__gte=fecha_limite)
            .values("proforma__fecha", "proforma__id", "proforma__cliente__name", "cantidad", "precio_venta", "subtotal")
            .order_by("-proforma__fecha")
        )
        if selected_warehouse:
            ventas = ventas.filter(proforma__warehouse=selected_warehouse)
        producto = Producto.objects.get(id=producto_id)
    else:
        producto = None

    return render(request, "inv/reports/historial_ventas.html", {
        "productos": productos,
        "ventas": ventas,
        "producto": producto,
        "dias": dias,
        "title": "Historial de producto",
        "warehouses": warehouses,
        "selected_warehouse": selected_warehouse,
    })
    
def buscar_productos(request):
    query = request.GET.get('q', '')
    page = int(request.GET.get('page', 1))

    productos = Producto.objects.filter(
        Q(nombre__icontains=query) | Q(referencia_cruzada__icontains=query)
    ).order_by('nombre')

    paginator = Paginator(productos, 10)  # 10 productos por página
    productos_pagina = paginator.get_page(page)

    data = {
        "results": [{"id": p.id, "nombre": p.nombre, "description": p.descripcion } for p in productos_pagina],
        "has_next": productos_pagina.has_next()
    }
    return JsonResponse(data)

@login_required
def reporte_inventario(request):
    warehouses = accessible_warehouses(request.user)
    selected_warehouse = _selected_warehouse_for_user(request)

    if selected_warehouse:
        stock_records = (
            ProductStock.objects.filter(warehouse=selected_warehouse, quantity__gt=0)
            .select_related('product', 'product__brand')
            .order_by('location', 'product__nombre')
        )
        productos = []
        for stock_record in stock_records:
            product = stock_record.product
            product.report_stock = stock_record.quantity
            product.report_location = stock_record.location
            productos.append(product)
    else:
        productos = list(Producto.objects.filter(stock__gt=0).select_related('brand').order_by('location'))
        for product in productos:
            product.report_stock = product.stock
            product.report_location = product.location

    total_cost = sum((p.cost or 0) * p.report_stock for p in productos)
    total_price = sum((p.precio or 0) * p.report_stock for p in productos)
    total_productos = len(productos)

    # Resumen por marca
    brand_summary = {}
    for p in productos:
        brand_name = p.brand.name if p.brand else "Sin Marca"
        if brand_name not in brand_summary:
            brand_summary[brand_name] = {"count": 0, "stock": 0, "cost_value": 0, "price_value": 0}
        brand_summary[brand_name]["count"] += 1
        brand_summary[brand_name]["stock"] += p.report_stock
        brand_summary[brand_name]["cost_value"] += float(p.cost or 0) * p.report_stock
        brand_summary[brand_name]["price_value"] += float(p.precio or 0) * p.report_stock
    brand_summary = sorted(brand_summary.items(), key=lambda x: x[1]["stock"], reverse=True)

    # Resumen por ubicación
    location_summary = {}
    for p in productos:
        loc = p.report_location or "Sin Ubicación"
        if loc not in location_summary:
            location_summary[loc] = {"count": 0, "stock": 0}
        location_summary[loc]["count"] += 1
        location_summary[loc]["stock"] += p.report_stock
    location_summary = sorted(location_summary.items(), key=lambda x: x[1]["stock"], reverse=True)

    # Exportar a Excel
    if request.GET.get("export") == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        ws.append(["ID", "Nombre", "Referencia Cruzada", "Marca", "Stock", "Ubicación", "Costo", "Precio"])
        for p in productos:
            ws.append([
                p.id,
                p.nombre,
                p.referencia_cruzada or "",
                p.brand.name if p.brand else "",
                p.report_stock,
                p.report_location or "",
                float(p.cost or 0),
                float(p.precio or 0),
            ])
        # Fila de totales
        ws.append(["", "TOTALES", "", "", sum(p.report_stock for p in productos), "",
                   float(total_cost), float(total_price)])

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_inventario.xlsx"'
        return response

    context = {
        "title": "Reporte de Inventario",
        "total_cost": total_cost,
        "total_price": total_price,
        "total_productos": total_productos,
        "brand_summary": brand_summary,
        "location_summary": location_summary,
        "warehouses": warehouses,
        "selected_warehouse": selected_warehouse,
    }

    return render(request, "inv/reports/reporte_inventario.html", context)
    
def proforma_report(request):
    proformas = Proforma.objects.none()
    total_general = 0
    print_mode = request.GET.get('print') == '1'

    today = timezone.localdate()
    default_start = today.replace(day=1)

    fecha_inicio = request.GET.get('fecha_inicio') or default_start.strftime('%Y-%m-%d')
    fecha_fin = request.GET.get('fecha_fin') or today.strftime('%Y-%m-%d')

    month_labels = []
    month_amounts = []
    seller_labels = []
    seller_sales = []

    try:
        fi_date = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        ff_date = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Formato de fecha invalido. Se aplico el rango por defecto del mes actual.')
        fi_date = default_start
        ff_date = today
        fecha_inicio = fi_date.strftime('%Y-%m-%d')
        fecha_fin = ff_date.strftime('%Y-%m-%d')

    if fi_date < ff_date:
        fi = make_aware(datetime.combine(fi_date, time.min))
        ff = make_aware(datetime.combine(ff_date + timedelta(days=1), time.min))

        base_queryset = Proforma.objects.filter(
            estado='EJECUTADO',
            fecha__gte=fi,
            fecha__lt=ff,
        )

        proformas_queryset = base_queryset.order_by('-fecha')

        # Calcular total general antes de paginar
        total_general = sum(p.total_neto() for p in proformas_queryset)

        if print_mode:
            # En modo impresión se muestra toda la tabla sin paginación.
            proformas = proformas_queryset
        else:
            # Paginación para navegación en pantalla.
            paginator = Paginator(proformas_queryset, 10)  # 10 por página
            page_number = request.GET.get('page')
            proformas = paginator.get_page(page_number)

        daily_amount_map = {}
        seller_totals = {}
        for proforma in base_queryset.select_related('usuario'):
            day_date = timezone.localtime(proforma.fecha).date()
            daily_amount_map[day_date] = daily_amount_map.get(day_date, 0.0) + float(proforma.total_neto())

            seller_name = getattr(proforma.usuario, 'name', None) or getattr(proforma.usuario, 'username', None) or 'Sin vendedor'
            seller_totals[seller_name] = seller_totals.get(seller_name, 0.0) + float(proforma.total_neto())

        current_day = fi_date
        while current_day <= ff_date:
            month_labels.append(current_day.strftime('%d/%m'))
            month_amounts.append(round(daily_amount_map.get(current_day, 0.0), 2))
            current_day += timedelta(days=1)

        seller_labels = list(seller_totals.keys())
        seller_sales = [round(value, 2) for value in seller_totals.values()]
    else:
        messages.error(request, 'La fecha inicio debe ser menor que la fecha fin.')

    context = {
        'proformas': proformas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'print_mode': print_mode,
        'total_general': total_general,
        'title': 'Reporte de Proformas',
        'month_labels_json': json.dumps(month_labels),
        'month_amounts_json': json.dumps(month_amounts),
        'seller_labels_json': json.dumps(seller_labels),
        'seller_sales_json': json.dumps(seller_sales),
    }
    return render(request, 'inv/reports/proforma_report.html', context)

# COMPRAS

@login_required
def purchase_list(request):
    query = request.GET.get('q')
    tipo = request.GET.get('tipo_busqueda', 'id')
    purchases = _accessible_purchases(request.user).order_by('-id', '-date', 'status')
    if query:
        if tipo == 'id':
            purchases = purchases.filter(id__icontains=query)
        elif tipo == 'proveedor':
            purchases = purchases.filter(supplier__name__icontains=query)
        elif tipo == 'factura':
            purchases = purchases.filter(invoice_number__icontains=query)
    paginator = Paginator(purchases, 10)
    page_number = request.GET.get('page')
    purchases = paginator.get_page(page_number)
    context = {
        'purchases': purchases,
        'title': 'Lista de Compras',
        'subtitle': 'Lista de compras registradas',
        'icon': 'fa-shopping-cart',
        'tipo_busqueda': tipo,  # Para mantener el select en el template
        'q': query, # Para mantener el valor de búsqueda en el template
    }
    return render(request, 'inv/purchase/purchase_list.html', context)

@login_required(login_url='login')
def create_purchase(request):
    company = getattr(request.user, 'company', None)
    default_sale_margin_percentage = float(getattr(company, 'default_sale_margin_percentage', 35) or 35)
        
    if request.method == 'POST':
        form = PurchaseForm(request.POST, user=request.user)

        # Primero validar el formulario principal para construir la instancia
        if form.is_valid():
            purchase = form.save(commit=False)
            purchase.user = request.user

            # Asociar el formset a la instancia (aunque no esté guardada aún)
            formset = PurchaseDetailFormSet(request.POST, instance=purchase)

            if formset.is_valid():
                with transaction.atomic():
                    # Guardar purchase y detalles
                    purchase.save()
                    formset.instance = purchase
                    formset.save()

                    # Calcular el total de la compra
                    total = 0
                    for f in formset.forms:
                        if f.cleaned_data and not f.cleaned_data.get('DELETE', False):
                            qty = f.cleaned_data.get('quantity', 0)
                            price = f.cleaned_data.get('unit_price', 0)
                            total += qty * price

                    purchase.total_amount = total
                    purchase.save()

                    if purchase.status == 'confirmed':
                        confirm_purchase_and_apply_inventory(purchase, user=request.user)
                        messages.success(request, "Compra confirmada correctamente.")
                        return redirect('purchase_list')
                    else:
                        messages.success(request, "Compra registrada correctamente.")
                        return redirect('update_purchase', pk=purchase.pk)
            else:
                # formset inválido: renderizar con errores
                messages.error(request, "Error en los detalles de la compra.")
        else:
            # form inválido: para renderizar la página con los datos POST
            formset = PurchaseDetailFormSet(request.POST)
    else:
        form = PurchaseForm(user=request.user)
        formset = PurchaseDetailFormSet()

    return render(request, 'inv/purchase/create_purchase.html', {
        'form': form,
        'formset': formset,
        'title': 'Registrar Compra',
        'default_sale_margin_percentage': default_sale_margin_percentage,
    })

@login_required(login_url='login')
def update_purchase(request, pk):
    purchase = get_object_or_404(_accessible_purchases(request.user), pk=pk)
    company = getattr(request.user, 'company', None)
    default_sale_margin_percentage = float(getattr(company, 'default_sale_margin_percentage', 35) or 35)
    
    if purchase.status == 'confirmed':
        messages.warning(request, "Esta compra ya está confirmada y no se puede modificar.")
        return redirect('purchase_list') 
    
    if request.method == 'POST':
        form = PurchaseForm(request.POST, instance=purchase, user=request.user)
        formset = PurchaseDetailFormSet(request.POST, instance=purchase)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                purchase = form.save(commit=False)
                purchase.user = request.user
                #purchase.date = now()
                purchase.save()  # Guardar primero para poder asignarlo a los detalles

                total = 0
                for form in formset.forms:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                        detail = form.save(commit=False)
                        detail.purchase = purchase
                        detail.save()
                        total += detail.quantity * detail.unit_price

                formset.save()  # 🔥 Aquí se eliminan los marcados con DELETE

                purchase.total_amount = total
                purchase.save()
                
                if purchase.status == 'confirmed':
                    # Redirigir a la lista de compras
                    confirm_purchase_and_apply_inventory(purchase, user=request.user)
                    
                    messages.success(request, "Compra confirmada y actualizada correctamente.")
                    return redirect('purchase_list')  
                else:   
                    # Si no está confirmado, redirigir a la misma página de actualización
                    messages.success(request, "Compra actualizada correctamente.")
                    return redirect('update_purchase', pk=purchase.pk)
        else:
            # Agregar errores del formulario principal
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form.fields[field].label if field in form.fields else field
                    messages.error(request, f"{field_label}: {error}")
            
            # Agregar errores del formset
            for i, formset_error in enumerate(formset.errors):
                if formset_error:
                    for field, errors in formset_error.items():
                        for error in errors:
                            messages.error(request, f"Producto {i+1} - {field}: {error}")
            
            # Mensaje genérico al final
            if form.errors or formset.errors:
                messages.error(request, "Por favor, corrija los errores señalados.")
    else:
        form = PurchaseForm(instance=purchase, user=request.user)
        formset = PurchaseDetailFormSet(instance=purchase)

    details_with_subtotals = [
        {'form': form, 'subtotal': form.instance.subtotal() if form.instance.pk else 0}
        for form in formset
    ]

    return render(request, 'inv/purchase/create_purchase.html', {
        'form': form,
        'formset': formset,
        'purchase': purchase,
        'details': details_with_subtotals,
        'title': 'Actualizar Compra',
        'default_sale_margin_percentage': default_sale_margin_percentage,
    })

def revert_purchase_movement(purchase):
    ct = ContentType.objects.get_for_model(purchase)

    # Evita duplicar egresos si la anulación se procesa más de una vez.
    existing_out = Movement.objects.filter(
        content_type=ct,
        object_id=purchase.id,
        movement_type='OUT'
    ).first()
    if existing_out:
        return existing_out

    movement_in = purchase.movement
    if not movement_in:
        return None
    if movement_in.warehouse_id is None:
        raise ValueError('El movimiento de compra no tiene almacén asignado.')

    movement_items = list(movement_in.items.select_related('product'))
    insufficient = []
    for item in movement_items:
        warehouse_stock = ProductStock.objects.filter(
            product=item.product,
            warehouse=movement_in.warehouse,
        ).first()
        current_stock = warehouse_stock.quantity if warehouse_stock else 0
        if current_stock < item.quantity:
            insufficient.append(
                f"{item.product.nombre} (stock actual: {current_stock}, requiere: {item.quantity})"
            )

    if insufficient:
        raise ValueError("Stock insuficiente para anular la compra: " + "; ".join(insufficient))

    # Crear movimiento de egreso solo si el stock alcanza para todos los items.
    movement_out = Movement.objects.create(
        movement_type='OUT',
        warehouse=movement_in.warehouse,
        content_type=ct,
        object_id=purchase.id,
        user=purchase.user,
        description=f"Egreso por anulación de compra #{purchase.id}"
    )
    for item in movement_items:
        MovementItem.objects.create(
            movement=movement_out,
            product=item.product,
            quantity=item.quantity,  # misma cantidad que el ingreso
            unit_price=item.unit_price or item.product.cost  # Usar el costo guardado en el movimiento original
        )
        apply_warehouse_stock_change(item.product, movement_in.warehouse, -item.quantity)
    return movement_out

# No se puede eliminar una compra, solo se puede anular
@login_required
def cancelled_purchase(request, pk):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                purchase = get_object_or_404(_accessible_purchases(request.user).select_for_update(), pk=pk)

                if purchase.status == 'cancelled':
                    messages.warning(request, "La compra ya estaba anulada.")
                    return redirect('purchase_list')

                revert_purchase_movement(purchase)
                purchase.status = 'cancelled'
                purchase.save(update_fields=['status'])

            messages.success(request, "Compra anulada y stock revertido correctamente.")
            return redirect('purchase_list')
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect('purchase_list')

    purchase = get_object_or_404(_accessible_purchases(request.user), pk=pk)
    return render(request, 'inv/purchase/cancelled_purchase.html', {
        'purchase': purchase,
    })

@login_required(login_url='login')
def delete_purchase_detail(request, pk):
    purchase_detail = get_object_or_404(PurchaseDetail.objects.filter(purchase__in=_accessible_purchases(request.user)), pk=pk)
    if request.method == 'POST':
        purchase_detail.delete()
        messages.success(request, "Detalle de compra eliminado correctamente.")
        return redirect('update_purchase', pk=purchase_detail.purchase.pk)  # Cambia por tu URL real
    return render(request, 'inv/purchase/delete_purchase_detail.html', {
        'purchase_detail': purchase_detail,
    })

@login_required(login_url='login')
def purchase_detail(request, pk):
    purchase = get_object_or_404(_accessible_purchases(request.user), pk=pk)
    details = PurchaseDetail.objects.filter(purchase=purchase)
    print_mode = request.GET.get('print') == '1'
    context = {
        'purchase': purchase,
        'details': details,
        'title': 'Compra',
        'subtitle': 'Detalles de la compra',
        'icon': 'fa-shopping-cart',
        'print_mode': print_mode,
    }
    
    return render(request, 'inv/purchase/purchase.html', context)

def create_purchase_movement(purchase):
    # Compatibilidad temporal: delega al servicio unico para evitar rutas duplicadas.
    return confirm_purchase_and_apply_inventory(purchase, user=getattr(purchase, 'user', None))

# MOVIMIENTOS DE INVENTARIO
@login_required
def movement_list(request):
    product_id = request.GET.get('producto_id')
    movements = _accessible_movements(request.user).select_related('warehouse').prefetch_related('items__product').order_by('-id', '-date')
    selected_producto_nombre = None
    if product_id:
        try:
            movements = movements.filter(items__product_id=product_id).distinct()
            selected_producto = Producto.objects.get(id=product_id)
            selected_producto_nombre = selected_producto.nombre + " - " + selected_producto.descripcion
        except Producto.DoesNotExist:
            messages.error(request, "Producto no encontrado.")
            return redirect('movement_list')
    paginator = Paginator(movements, 10)  # 10 movimientos por página
    page_number = request.GET.get('page')
    movements = paginator.get_page(page_number)
    context = {
        'movements': movements,
        'title': 'Movimientos',
        'subtitle': 'Lista de movimientos',
        'icon': 'fa-exchange-alt',
        'selected_producto': product_id,
        'selected_producto_nombre': selected_producto_nombre,
    }   
    return render(request, 'inv/movement/movement_list.html', context)

@login_required(login_url='login')
def movement_detail(request, pk):
    # Obtener el movimiento usando el ID
    movement = get_object_or_404(_accessible_movements(request.user), id=pk)

    # Obtener los MovementItems relacionados con el movimiento
    movement_items = movement.items.all()

    # Retornar la plantilla con el movimiento y sus items
    return render(request, 'inv/movement/movement_detail.html', {
        'movement': movement,
        'movement_items': movement_items,
    })
    
@login_required(login_url='login')
def create_movement(request):
    if request.method == 'POST':
        form = MovementForm(request.POST, user=request.user)
        formset = MovementItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                movement = form.save(commit=False)
                movement.user = request.user
                movement.date = now()
                movement.save()

                formset.instance = movement
                items = formset.save(commit=False)

                for item_form in formset:
                    if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                        product = item_form.cleaned_data.get('product')
                        quantity = item_form.cleaned_data.get('quantity')

                        # Calcular unit_price y subtotal
                        unit_price = product.cost
                        subtotal = unit_price * quantity

                        # Stock antes del movimiento
                        previous_stock = product.stock

                        delta = quantity if movement.movement_type == 'IN' else -quantity
                        warehouse_stock = apply_warehouse_stock_change(
                            product,
                            movement.warehouse,
                            delta,
                        )

                        # Completar los datos del MovementItem
                        item = item_form.save(commit=False)
                        item.unit_price = unit_price
                        item.subtotal = subtotal
                        item.stock_after_movement = warehouse_stock.quantity
                        item.save()

                # Guardar eliminados (si se usó `can_delete`)
                formset.save_m2m()

                messages.success(request, "Movimiento registrado correctamente.")
                return redirect('movement_list')
        else:
            messages.error(request, "Error al registrar el movimiento.")
    else:
        form = MovementForm(user=request.user)
        formset = MovementItemFormSet(queryset=MovementItem.objects.none())

    return render(request, 'inv/movement/movement_form.html', {
        'form': form,
        'formset': formset,
        'warehouses': accessible_warehouses(request.user),
        'default_warehouse': default_user_warehouse(request.user),
    })

def get_producto(request, id):
    producto = Producto.objects.get(pk=id)
    return JsonResponse({
        'id': producto.id,
        'nombre': producto.nombre,
        'cost': float(producto.cost)
    })

@login_required(login_url='login')
def movement_pdf(request, pk):
    movement = get_object_or_404(_accessible_movements(request.user), pk=pk)
    html_string = render_to_string('inv/movement/pdf.html', {
        'movement': movement,
        'items': movement.items.all()
    })

    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = BytesIO()
    html.write_pdf(pdf_file)
    pdf_file.seek(0)

    response = HttpResponse(pdf_file.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'filename=movement_{movement.id}.pdf'

    return response


@login_required
def cargar_inventario_inicial(request):
    from core.services.price_evaluation_service import PriceEvaluationService
    company = getattr(request.user, 'company', None)
    if not company or not getattr(company, 'enable_initial_stock_load', False):
        messages.error(request, 'La carga inicial de stock está deshabilitada para tu empresa.')
        return redirect('movement_list')

    if request.method == 'POST':
        form = InventoryUploadForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            nombre_archivo = archivo.name.lower()
            update_descripcion = form.cleaned_data.get('actualizar_descripcion', False)
            update_costo = form.cleaned_data.get('actualizar_costo', False)
            update_precio = form.cleaned_data.get('actualizar_precio', False)
            update_ubicacion = form.cleaned_data.get('actualizar_ubicacion', False)
            errores = []
            items_a_crear = []
            if nombre_archivo.endswith('.xlsx'):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    product_code = str(row[0]).strip() if row[0] else None
                    quantity = int(row[1]) if row[1] is not None else 0
                    cost = float(row[2]) if len(row) > 2 and row[2] is not None else None
                    precio = float(row[3]) if len(row) > 3 and row[3] is not None else None
                    location = str(row[4]).strip() if len(row) > 4 and row[4] else None
                    descripcion_producto = str(row[5]).strip() if len(row) > 5 and row[5] else None
                    if not product_code or quantity <= 0:
                        errores.append(f"Línea {i}: Código de producto y cantidad son obligatorios.")
                        continue
                    producto = Producto.objects.filter(nombre=product_code).first()
                    if not producto:
                        # Crear producto si no existe
                        producto = Producto.objects.create(
                            nombre=product_code,
                            descripcion=descripcion_producto or "",
                            cost=cost or 0,
                            precio=precio or 0,
                            stock=0,
                            location=location or ""
                        )
                        # Crear historial de precio inicial
                        if precio is not None:
                            PriceEvaluationService.propose_new_price(
                                product=producto,
                                old_price=0,
                                new_price=precio,
                                user=request.user,
                                reason="Carga de inventario inicial",
                                cost_reference=cost,
                                change_type='INICIAL',
                                auto_approve_on_increase=True
                            )
                        items_a_crear.append((producto, quantity, cost, precio, True))
                    else:
                        # Producto ya existe
                        crear_historial = False
                        precio_anterior = float(producto.precio or 0)

                        if update_descripcion and descripcion_producto is not None:
                            producto.descripcion = descripcion_producto

                        if update_costo and cost is not None:
                            producto.cost = cost

                        if update_ubicacion and location:
                            producto.location = location

                        if update_precio and precio is not None:
                            if producto.stock and precio > precio_anterior:
                                producto.precio = precio
                                crear_historial = True
                            elif not producto.stock:
                                producto.precio = precio
                                crear_historial = True

                        items_a_crear.append((producto, quantity, cost, precio, crear_historial))
                # Fin for
            else:
                errores.append("Solo se permiten archivos Excel (.xlsx).")

            if errores:
                for error in errores:
                    messages.error(request, error)
                return render(request, 'inv/movement/cargar_inventario.html', {'form': form})

            # Crear movimiento y cargar items
            with transaction.atomic():
                movimiento = Movement.objects.create(
                    movement_type='IN',
                    warehouse=form.cleaned_data['warehouse'],
                    description=form.cleaned_data.get('descripcion') or 'Inventario inicial',
                    user=request.user
                )
                for producto, quantity, cost, precio, crear_historial in items_a_crear:
                    MovementItem.objects.create(
                        movement=movimiento,
                        product=producto,
                        quantity=int(quantity),
                        unit_price=cost  # Guardar el costo con el que se creó el movimiento
                    )
                    apply_warehouse_stock_change(
                        producto,
                        movimiento.warehouse,
                        int(quantity),
                        location=producto.location or '',
                    )
                    # Crear historial de precio si corresponde
                    if crear_historial and precio is not None:
                        PriceEvaluationService.propose_new_price(
                            product=producto,
                            old_price=precio,
                            new_price=precio,
                            user=request.user,
                            reason="Actualización por inventario inicial",
                            cost_reference=cost,
                            change_type='INICIAL',
                            auto_approve_on_increase=True
                        )
            messages.success(request, "Inventario inicial cargado correctamente.")
            return redirect('movement_list')
    else:
        form = InventoryUploadForm(user=request.user)
    return render(request, 'inv/movement/cargar_inventario.html', {'form': form})

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')  # ⚠️ solo si no usas CSRF token con JS
class CreateMovementView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            if not data.get('warehouse_id'):
                raise ValueError('Debes seleccionar un almacén.')
            if not data.get('items'):
                raise ValueError('Debes agregar al menos un producto.')
            warehouse = resolve_user_warehouse(request.user, data['warehouse_id'])

            with transaction.atomic():
                # Crear movimiento principal
                movement = Movement.objects.create(
                    movement_type=data['movement_type'],
                    warehouse=warehouse,
                    description=data.get('description', ''),
                    user=request.user,
                    status='COMPLETED'
                )

                for item in data['items']:
                    producto = Producto.objects.get(id=item['product_id'])
                    cantidad = int(item['quantity'])

                    if cantidad <= 0:
                        raise ValueError('La cantidad debe ser mayor que cero.')

                    delta = cantidad if movement.movement_type == 'IN' else -cantidad
                    warehouse_stock = apply_warehouse_stock_change(producto, warehouse, delta)

                    # Crear ítem
                    MovementItem.objects.create(
                        movement=movement,
                        product=producto,
                        quantity=cantidad,
                        unit_price=producto.cost,
                        stock_after_movement=warehouse_stock.quantity,
                        observation=item.get('observation', '')
                    )

                # Retornar respuesta exitosa y direccionar a la lista de movimientos
            return JsonResponse({'message': 'Movimiento guardado correctamente.', 'movement_id': movement.id, 'redirect_url': reverse('movement_list')}, status=201 )



        except (Producto.DoesNotExist, Warehouse.DoesNotExist):
            return JsonResponse({'error': 'Producto o almacén no encontrado.'}, status=404)
        except (ValueError, WarehouseAccessDenied, InsufficientWarehouseStock) as ve:
            return JsonResponse({'error': str(ve)}, status=400)
        except Exception as e:
            return JsonResponse({'error': 'Ocurrió un error inesperado.', 'detalle': str(e)}, status=500)


@login_required(login_url='login')
def create_transfer(request):
    if not _can_manage_warehouses(request.user):
        messages.error(request, 'No tienes permisos para realizar transferencias.')
        return redirect('home')

    return render(request, 'inv/transfer/transfer_form.html', {
        'title': 'Transferencia entre almacenes',
        'warehouses': accessible_warehouses(request.user),
        'default_warehouse': default_user_warehouse(request.user),
    })


@method_decorator(login_required, name='dispatch')
class CreateTransferView(View):
    def post(self, request, *args, **kwargs):
        if not _can_manage_warehouses(request.user):
            return JsonResponse({'error': 'No tienes permisos para realizar transferencias.'}, status=403)

        try:
            data = json.loads(request.body)
            if not data.get('origin_warehouse_id') or not data.get('destination_warehouse_id'):
                raise ValueError('Debes seleccionar almacén de origen y destino.')
            if not data.get('items'):
                raise ValueError('Debes agregar al menos un producto.')

            origin_warehouse = resolve_user_warehouse(request.user, data['origin_warehouse_id'])
            destination_warehouse = resolve_user_warehouse(request.user, data['destination_warehouse_id'])
            product_ids = [int(item.get('product_id')) for item in data['items']]
            products = Producto.objects.in_bulk(product_ids)
            if len(products) != len(set(product_ids)):
                raise Producto.DoesNotExist

            items = [
                {
                    'product': products[product_id],
                    'quantity': item.get('quantity'),
                    'observation': item.get('observation', ''),
                }
                for item, product_id in zip(data['items'], product_ids)
            ]
            transfer = create_warehouse_transfer(
                origin_warehouse=origin_warehouse,
                destination_warehouse=destination_warehouse,
                items=items,
                user=request.user,
                description=data.get('description', ''),
            )
            return JsonResponse({
                'message': 'Transferencia registrada correctamente.',
                'transfer_id': transfer.id,
                'redirect_url': reverse('movement_list'),
            }, status=201)
        except (Producto.DoesNotExist, Warehouse.DoesNotExist):
            return JsonResponse({'error': 'Producto o almacén no encontrado.'}, status=404)
        except (TypeError, ValueError, WarehouseAccessDenied, InsufficientWarehouseStock) as exc:
            return JsonResponse({'error': str(exc)}, status=400)
        except Exception as exc:
            return JsonResponse({'error': 'Ocurrió un error inesperado.', 'detalle': str(exc)}, status=500)


# PRE-INVENTARIO

@login_required(login_url='login')
def pre_inventario(request):
    ubicacion = request.GET.get('ubicacion', '__all__')
    query = request.GET.get('q', '')
    con_stock = request.GET.get('con_stock') == 'on'
    sin_costo = request.GET.get('sin_costo') == 'on'
    sin_precio = request.GET.get('sin_precio') == 'on'
    warehouses = accessible_warehouses(request.user)
    selected_warehouse = _selected_warehouse_for_user(request)

    if selected_warehouse:
        stocks = ProductStock.objects.filter(warehouse=selected_warehouse).select_related('product', 'product__brand')
        if ubicacion == "":
            stocks = stocks.filter(Q(location__exact="") | Q(location__isnull=True))
        elif ubicacion != "__all__":
            stocks = stocks.filter(location__iexact=ubicacion)
        if query:
            stocks = stocks.filter(
                Q(product__nombre__icontains=query)
                | Q(product__referencia_cruzada__icontains=query)
                | Q(product__descripcion__icontains=query)
            )
        if con_stock:
            stocks = stocks.filter(quantity__gt=0)
        if sin_costo:
            stocks = stocks.filter(Q(product__cost__isnull=True) | Q(product__cost=0))
        if sin_precio:
            stocks = stocks.filter(Q(product__precio__isnull=True) | Q(product__precio=0))

        productos = []
        for stock in stocks.order_by('location', 'product__nombre'):
            product = stock.product
            product.report_stock = stock.quantity
            product.report_location = stock.location
            productos.append(product)
        ubicaciones = ProductStock.objects.filter(warehouse=selected_warehouse).exclude(location__exact="").values_list('location', flat=True).distinct().order_by('location')
    else:
        productos = Producto.objects.all()

    # Filtro de ubicación
    if not selected_warehouse and ubicacion == "":
        # Solo productos sin ubicación (location vacío o None)
        productos = productos.filter(Q(location__exact="") | Q(location__isnull=True))
    elif not selected_warehouse and ubicacion != "__all__":
        productos = productos.filter(location__iexact=ubicacion)
    # Si es "__all__", no se filtra por ubicación

    # Filtro de búsqueda
    if not selected_warehouse and query:
        productos = productos.filter(
            Q(nombre__icontains=query)
            | Q(referencia_cruzada__icontains=query)
            | Q(descripcion__icontains=query)
        )
    # Filtro de stock
    if not selected_warehouse and con_stock:
        productos = productos.filter(stock__gt=0)

    # Filtro sin costo
    if not selected_warehouse and sin_costo:
        productos = productos.filter(Q(cost__isnull=True) | Q(cost=0))
    
    # Filtro sin precio
    if not selected_warehouse and sin_precio:
        productos = productos.filter(Q(precio__isnull=True) | Q(precio=0))
        
        
    # Lista de ubicaciones distintas (sin None ni vacío)
    if not selected_warehouse:
        ubicaciones = Producto.objects.exclude(location__isnull=True).exclude(location__exact="").values_list('location', flat=True).distinct().order_by('location')

    if not selected_warehouse:
        productos = list(productos)
        for product in productos:
            product.report_stock = product.stock
            product.report_location = product.location

    paginator = Paginator(productos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'productos': page_obj,
        'ubicaciones': ubicaciones,
        'ubicacion_actual': ubicacion,
        'query': query,
        'con_stock': con_stock,
        'sin_costo': sin_costo,
        'sin_precio': sin_precio,
        'warehouses': warehouses,
        'selected_warehouse': selected_warehouse,
        'title': 'Pre-inventario por ubicación',
        'placeholder': 'Buscar por código, referencia cruzada o descripción'
    }
    return render(request, 'inv/reports/pre_inventario.html', context)