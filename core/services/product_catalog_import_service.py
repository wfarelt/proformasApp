import hashlib
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import requests

from django.db import transaction
from django.db.models import F
from django.db.models.functions import Trim, Upper

from openpyxl import load_workbook
from openpyxl import Workbook

from core.models import Producto

# URL pública del index.json en el repositorio de catálogos
CLOUD_INDEX_URL = "https://raw.githubusercontent.com/wfarelt/proformasApp/master/catalog_cloud/index.json"

# Dominios permitidos para descargar archivos de catálogo
ALLOWED_DOWNLOAD_DOMAINS = {"raw.githubusercontent.com"}

# Tamaño máximo permitido por descarga (10 MB)
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024


@dataclass
class CatalogImportRow:
    row_number: int
    codigo: str
    descripcion: str


class ProductCatalogImportService:
    REQUIRED_COLUMNS = {
        "codigo": ["codigo", "código", "code", "sku", "nombre"],
        "descripcion": ["descripcion", "descripción", "description"],
    }

    @classmethod
    def build_template_file(cls) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "catalogo"

        sheet.append(["Código", "Descripción"])
        sheet.append(["PROD-001", "Producto de ejemplo"])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @classmethod
    def import_from_excel(cls, uploaded_file) -> Dict:
        workbook = load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active

        header_map = cls._build_header_map(sheet)
        rows, row_errors = cls._parse_rows(sheet, header_map)

        if not rows:
            return {
                "created": 0,
                "skipped_existing": 0,
                "duplicate_in_file": 0,
                "errors": row_errors or ["No se encontraron filas válidas para importar."],
                "total_rows": 0,
            }

        imported_codes = [row.codigo for row in rows]
        duplicated_in_file = cls._find_duplicates(imported_codes)

        existing_codes = cls._get_existing_codes(imported_codes)

        rows_to_create: List[CatalogImportRow] = []
        skipped_existing = 0

        for row in rows:
            normalized_code = cls._normalize_text(row.codigo)
            if normalized_code in duplicated_in_file:
                continue
            if normalized_code in existing_codes:
                skipped_existing += 1
                continue
            rows_to_create.append(row)

        created = cls._bulk_create_catalog(rows_to_create)

        return {
            "created": created,
            "skipped_existing": skipped_existing,
            "duplicate_in_file": len(duplicated_in_file),
            "errors": row_errors,
            "total_rows": len(rows),
        }

    @classmethod
    def _build_header_map(cls, sheet) -> Dict[str, int]:
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not raw_headers:
            raise ValueError("El archivo Excel no tiene encabezados.")

        normalized_headers = {
            cls._normalize_text(str(value)): index
            for index, value in enumerate(raw_headers)
            if value is not None and str(value).strip()
        }

        def find_column(possible_names: List[str]) -> Optional[int]:
            for name in possible_names:
                normalized_name = cls._normalize_text(name)
                if normalized_name in normalized_headers:
                    return normalized_headers[normalized_name]
            return None

        codigo_col = find_column(cls.REQUIRED_COLUMNS["codigo"])
        if codigo_col is None:
            raise ValueError("No se encontró la columna obligatoria 'codigo'.")

        descripcion_col = find_column(cls.REQUIRED_COLUMNS["descripcion"])
        if descripcion_col is None:
            raise ValueError("No se encontró la columna obligatoria 'descripcion'.")

        return {
            "codigo": codigo_col,
            "descripcion": descripcion_col,
        }

    @classmethod
    def _parse_rows(cls, sheet, header_map: Dict[str, Optional[int]]) -> Tuple[List[CatalogImportRow], List[str]]:
        rows: List[CatalogImportRow] = []
        errors: List[str] = []

        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            codigo = cls._cell_value(values, header_map["codigo"])
            if not codigo:
                continue

            codigo = cls._normalize_text(codigo)
            descripcion = cls._cell_value(values, header_map.get("descripcion"))

            if len(codigo) > 100:
                errors.append(f"Fila {row_index}: código supera 100 caracteres.")
                continue

            if not descripcion:
                errors.append(f"Fila {row_index}: descripción vacía.")
                continue

            rows.append(
                CatalogImportRow(
                    row_number=row_index,
                    codigo=codigo,
                    descripcion=descripcion,
                )
            )

        return rows, errors

    @classmethod
    def _bulk_create_catalog(cls, rows: List[CatalogImportRow]) -> int:
        if not rows:
            return 0

        products_to_create = []

        for row in rows:
            products_to_create.append(
                Producto(
                    nombre=row.codigo,
                    descripcion=row.descripcion,
                    stock=0,
                    cost=0,
                    precio=0,
                    latest_price=0,
                )
            )

        with transaction.atomic():
            created = Producto.objects.bulk_create(products_to_create, batch_size=1000)

        return len(created)

    @classmethod
    def _get_existing_codes(cls, codes: List[str]) -> set:
        normalized_codes = {cls._normalize_text(code) for code in codes}
        if not normalized_codes:
            return set()

        existing_qs = Producto.objects.annotate(normalized_code=Upper(Trim(F("nombre")))).filter(
            normalized_code__in=normalized_codes
        ).values_list("normalized_code", flat=True)

        return {cls._normalize_text(value) for value in existing_qs}

    @staticmethod
    def _find_duplicates(codes: List[str]) -> set:
        seen = set()
        duplicates = set()
        for code in codes:
            if code in seen:
                duplicates.add(code)
                continue
            seen.add(code)
        return duplicates

    @staticmethod
    def _cell_value(values, index: Optional[int]) -> str:
        if index is None:
            return ""
        value = values[index] if index < len(values) else None
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _normalize_text(text: str) -> str:
        return str(text).strip().upper() if text is not None else ""

    # ------------------------------------------------------------------ #
    # Métodos para importación desde catálogo en la nube                  #
    # ------------------------------------------------------------------ #

    @classmethod
    def fetch_cloud_index(cls) -> List[Dict]:
        """Descarga y valida el index.json del repositorio de catálogos en la nube.
        Retorna la lista de catálogos disponibles."""
        try:
            response = requests.get(CLOUD_INDEX_URL, timeout=10)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise ValueError(f"No se pudo obtener el índice de catálogos: {exc}")

        try:
            data = response.json()
        except ValueError:
            raise ValueError("El índice de catálogos tiene un formato inválido.")

        if not isinstance(data, dict) or "catalogs" not in data:
            raise ValueError("El índice de catálogos no tiene la estructura esperada.")

        return data.get("catalogs", [])

    @classmethod
    def import_from_cloud_url(cls, url: str, expected_checksum: str) -> Dict:
        """Descarga un archivo xlsx desde una URL confiable, verifica su checksum
        y ejecuta la importación reutilizando el pipeline existente."""
        cls._validate_download_url(url)

        try:
            response = requests.get(url, timeout=30, stream=True)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise ValueError(f"No se pudo descargar el archivo: {exc}")

        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            if len(content) > MAX_FILE_SIZE_BYTES:
                raise ValueError("El archivo supera el tamaño máximo permitido (10 MB).")

        cls._verify_checksum(content, expected_checksum)

        return cls.import_from_excel(BytesIO(content))

    @staticmethod
    def _validate_download_url(url: str) -> None:
        """Verifica que la URL pertenezca a un dominio confiable."""
        parsed = urlparse(url)
        if parsed.scheme != "https":
            raise ValueError("Solo se permiten URLs con HTTPS.")
        if parsed.netloc not in ALLOWED_DOWNLOAD_DOMAINS:
            raise ValueError(f"Dominio no permitido: '{parsed.netloc}'.")

    @staticmethod
    def _verify_checksum(content: bytes, expected: str) -> None:
        """Verifica el checksum SHA-256 del contenido descargado."""
        if not expected or not expected.startswith("sha256:"):
            return  # Sin checksum declarado, se omite la verificación
        expected_hash = expected.removeprefix("sha256:")
        actual_hash = hashlib.sha256(content).hexdigest()
        if actual_hash != expected_hash:
            raise ValueError("El archivo descargado no coincide con el checksum esperado. Posible archivo corrupto o alterado.")
