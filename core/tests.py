from django.test import TestCase
from django.urls import reverse
from unittest.mock import patch
from io import BytesIO
from django.db import IntegrityError, transaction
from core.services.inventory_service import InsufficientWarehouseStock, apply_warehouse_stock_change
from core.services.warehouse_access_service import WarehouseAccessDenied, resolve_user_warehouse
from inv.services.warehouse_transfer_service import create_warehouse_transfer

from openpyxl import Workbook, load_workbook
from django.core.files.uploadedfile import SimpleUploadedFile

from core.models import Company, Detalle, Producto, ProductStock, Proforma, User, Warehouse
from core.services.product_catalog_import_service import ProductCatalogImportService


class ProformaRecommendationTests(TestCase):
	def setUp(self):
		self.company = Company.objects.create(
			name='Empresa Test',
			tax_id='123456',
			email='empresa@test.com',
			enable_product_recommendations=True,
		)
		self.user = User.objects.create_user(
			username='tester',
			email='tester@test.com',
			name='Tester',
			password='secret123',
			company=self.company,
		)
		self.client.force_login(self.user)

	def _create_producto(self, codigo, precio=10):
		return Producto.objects.create(nombre=codigo, precio=precio, latest_price=precio, stock=10)

	def _create_proforma(self, estado='PENDIENTE', cliente=None):
		return Proforma.objects.create(
			usuario=self.user,
			company=self.company,
			estado=estado,
			cliente=cliente,
		)

	def _add_detalle(self, proforma, producto, cantidad=1, precio=None):
		precio = precio or producto.precio
		return Detalle.objects.create(
			proforma=proforma,
			producto=producto,
			cantidad=cantidad,
			precio_venta=precio,
			subtotal=precio * cantidad,
		)

	def test_recommended_products_use_executed_first_then_pending_fallback(self):
		producto_a = self._create_producto('A-001')
		producto_b = self._create_producto('B-001')
		producto_c = self._create_producto('C-001')
		producto_d = self._create_producto('D-001')
		producto_e = self._create_producto('E-001')

		current_proforma = self._create_proforma()
		self._add_detalle(current_proforma, producto_c, cantidad=1)

		executed_proforma_1 = self._create_proforma(estado='EJECUTADO')
		self._add_detalle(executed_proforma_1, producto_c, cantidad=1)
		self._add_detalle(executed_proforma_1, producto_a, cantidad=2)
		self._add_detalle(executed_proforma_1, producto_b, cantidad=1)

		executed_proforma_2 = self._create_proforma(estado='EJECUTADO')
		self._add_detalle(executed_proforma_2, producto_c, cantidad=1)
		self._add_detalle(executed_proforma_2, producto_a, cantidad=1)

		executed_proforma_3 = self._create_proforma(estado='EJECUTADO')
		self._add_detalle(executed_proforma_3, producto_b, cantidad=1)

		pending_proforma = self._create_proforma(estado='PENDIENTE')
		self._add_detalle(pending_proforma, producto_c, cantidad=1)
		self._add_detalle(pending_proforma, producto_d, cantidad=5)

		anulado_proforma = self._create_proforma(estado='ANULADO')
		self._add_detalle(anulado_proforma, producto_c, cantidad=1)
		self._add_detalle(anulado_proforma, producto_e, cantidad=9)

		response = self.client.get(reverse('proforma_edit', args=[current_proforma.id]))

		recommended_products = response.context['recommended_products']
		recommended_ids = [producto.id for producto in recommended_products]

		self.assertEqual(recommended_ids, [producto_a.id, producto_b.id, producto_d.id])
		self.assertNotIn(producto_c.id, recommended_ids)
		self.assertNotIn(producto_e.id, recommended_ids)
		self.assertContains(response, 'Productos recomendados')

	def test_recommended_products_limit_pending_fallback_to_two(self):
		producto_base = self._create_producto('BASE-001')
		producto_p1 = self._create_producto('P-001')
		producto_p2 = self._create_producto('P-002')
		producto_p3 = self._create_producto('P-003')

		current_proforma = self._create_proforma()
		self._add_detalle(current_proforma, producto_base, cantidad=1)

		pending_1 = self._create_proforma(estado='PENDIENTE')
		self._add_detalle(pending_1, producto_base, cantidad=1)
		self._add_detalle(pending_1, producto_p1, cantidad=3)

		pending_2 = self._create_proforma(estado='PENDIENTE')
		self._add_detalle(pending_2, producto_base, cantidad=1)
		self._add_detalle(pending_2, producto_p2, cantidad=2)

		pending_3 = self._create_proforma(estado='PENDIENTE')
		self._add_detalle(pending_3, producto_base, cantidad=1)
		self._add_detalle(pending_3, producto_p3, cantidad=1)

		response = self.client.get(reverse('proforma_edit', args=[current_proforma.id]))

		recommended_ids = [producto.id for producto in response.context['recommended_products']]

		self.assertEqual(recommended_ids, [producto_p1.id, producto_p2.id])
		self.assertNotIn(producto_p3.id, recommended_ids)

	def test_recommended_products_can_be_disabled_per_company(self):
		self.company.enable_product_recommendations = False
		self.company.save(update_fields=['enable_product_recommendations'])

		producto = self._create_producto('A-001')
		executed_proforma = self._create_proforma(estado='EJECUTADO')
		self._add_detalle(executed_proforma, producto, cantidad=1)
		current_proforma = self._create_proforma()

		response = self.client.get(reverse('proforma_edit', args=[current_proforma.id]))

		self.assertFalse(response.context['enable_product_recommendations'])
		self.assertEqual(response.context['recommended_products'], [])
		self.assertNotContains(response, 'Productos recomendados')


class RoleAccessTests(TestCase):
	def setUp(self):
		self.company = Company.objects.create(
			name='Empresa Roles',
			tax_id='ROLE-123',
			email='roles@test.com',
		)

	def _create_user(self, username, role):
		return User.objects.create_user(
			username=username,
			email=f'{username}@test.com',
			name=username.title(),
			password='secret123',
			company=self.company,
			role=role,
		)

	def test_superadmin_sees_configuration_dashboard_only(self):
		user = self._create_user('superconfig', User.Roles.SUPERADMIN)
		self.client.force_login(user)

		response = self.client.get(reverse('home'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Panel de configuración')
		self.assertNotContains(response, 'Productos')

	def test_superadmin_is_redirected_from_operational_module(self):
		user = self._create_user('superblocked', User.Roles.SUPERADMIN)
		self.client.force_login(user)

		response = self.client.get(reverse('product_list'))

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('home'))

	def test_ventas_cannot_access_inventory_module(self):
		user = self._create_user('ventasuser', User.Roles.VENTAS)
		self.client.force_login(user)

		response = self.client.get(reverse('movement_list'))

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('home'))

	def test_almacen_can_access_inventory_module(self):
		user = self._create_user('almacenuser', User.Roles.ALMACEN)
		self.client.force_login(user)

		response = self.client.get(reverse('movement_list'))

		self.assertEqual(response.status_code, 200)

	def test_admin_can_open_user_management(self):
		user = self._create_user('adminpanel', User.Roles.ADMIN)
		self.client.force_login(user)

		response = self.client.get(reverse('user_list'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Usuarios')

	def test_admin_can_create_user_and_assign_role(self):
		user = self._create_user('admincreator', User.Roles.ADMIN)
		self.client.force_login(user)

		response = self.client.post(reverse('user_create'), {
			'username': 'nuevoalmacen',
			'email': 'nuevoalmacen@test.com',
			'name': 'Nuevo Almacen',
			'company': self.company.id,
			'role': User.Roles.ALMACEN,
			'password1': 'Secret12345*',
			'password2': 'Secret12345*',
		})

		self.assertEqual(response.status_code, 302)
		created_user = User.objects.get(username='nuevoalmacen')
		self.assertEqual(created_user.role, User.Roles.ALMACEN)

	def test_ventas_cannot_open_user_management(self):
		user = self._create_user('ventasblocked', User.Roles.VENTAS)
		self.client.force_login(user)

		response = self.client.get(reverse('user_list'))

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('home'))

	def test_admin_can_open_company_data_panel(self):
		user = self._create_user('admincompany', User.Roles.ADMIN)
		self.client.force_login(user)

		response = self.client.get(reverse('company_edit'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Datos de la empresa')

	def test_admin_can_update_company_general_data(self):
		user = self._create_user('adminupdatecompany', User.Roles.ADMIN)
		self.client.force_login(user)

		response = self.client.post(reverse('company_edit'), {
			'name': 'Empresa Actualizada',
			'tax_id': 'ROLE-123',
			'phone': '7777777',
			'email': 'empresa-actualizada@test.com',
			'address': 'Av. Principal 123',
			'city': 'La Paz',
			'website': 'https://empresa.test',
			'industry': 'Tecnología',
			'established_date': '2024-01-15',
		})

		self.assertEqual(response.status_code, 302)
		self.company.refresh_from_db()
		self.assertEqual(self.company.name, 'Empresa Actualizada')
		self.assertEqual(self.company.city, 'La Paz')

	def test_ventas_cannot_open_company_data_panel(self):
		user = self._create_user('ventascompanyblocked', User.Roles.VENTAS)
		self.client.force_login(user)

		response = self.client.get(reverse('company_edit'))

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('home'))


class CloudCatalogSuperadminTests(TestCase):
	def setUp(self):
		self.company = Company.objects.create(
			name='Empresa Catalogos',
			tax_id='CAT-123',
			email='catalogos@test.com',
		)
		self.superadmin = User.objects.create_user(
			username='supercatalog',
			email='supercatalog@test.com',
			name='Super Catalog',
			password='secret123',
			company=self.company,
			role=User.Roles.SUPERADMIN,
		)

	@patch('core.views.ProductCatalogImportService.publish_cloud_catalog_index_changes')
	@patch('core.views.ProductCatalogImportService.rename_cloud_catalog')
	def test_superadmin_can_rename_cloud_catalog(self, rename_mock, publish_mock):
		rename_mock.return_value = {
			'name': 'Catalogo Renombrado',
			'slug': 'electronica',
		}

		self.client.force_login(self.superadmin)
		response = self.client.post(reverse('superadmin_cloud_catalog_rename'), {
			'slug': 'electronica',
			'name': 'Catalogo Renombrado',
			'publish_now': 'on',
		})

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('superadmin_cloud_catalog_upload'))
		rename_mock.assert_called_once_with(slug='electronica', new_name='Catalogo Renombrado')
		publish_mock.assert_called_once_with(commit_message='Rename catalog electronica')

	@patch('core.views.ProductCatalogImportService.publish_cloud_catalog_delete')
	@patch('core.views.ProductCatalogImportService.delete_cloud_catalog')
	def test_superadmin_can_delete_cloud_catalog(self, delete_mock, publish_mock):
		delete_mock.return_value = {
			'catalog': {
				'name': 'Catalogo Legacy',
				'slug': 'legacy',
			},
			'deleted_file_path': None,
		}

		self.client.force_login(self.superadmin)
		response = self.client.post(reverse('superadmin_cloud_catalog_delete'), {
			'slug': 'legacy',
			'publish_now': 'on',
		})

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse('superadmin_cloud_catalog_upload'))
		delete_mock.assert_called_once_with(slug='legacy')
		publish_mock.assert_called_once_with(
			deleted_file_path=None,
			commit_message='Delete catalog legacy',
		)


class ProductCatalogImportTemplateTests(TestCase):
	def test_generated_template_contains_expected_headers(self):
		template_bytes = ProductCatalogImportService.build_template_file()
		workbook = load_workbook(BytesIO(template_bytes))
		sheet = workbook.active

		headers = [cell.value for cell in sheet[1]]
		self.assertEqual(headers, ['Código', 'Referencia cruzada', 'Descripción'])

	def test_import_maps_referencia_cruzada_and_descripcion(self):
		workbook = Workbook()
		sheet = workbook.active
		sheet.append(['Código', 'Referencia cruzada', 'Descripción'])
		sheet.append(['ABC-001', 'REF-001', 'Producto A'])

		buffer = BytesIO()
		workbook.save(buffer)
		buffer.seek(0)

		uploaded_file = SimpleUploadedFile(
			name='catalogo.xlsx',
			content=buffer.getvalue(),
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)

		result = ProductCatalogImportService.import_from_excel(uploaded_file)
		self.assertEqual(result['created'], 1)

		producto = Producto.objects.get(nombre='ABC-001')
		self.assertEqual(producto.referencia_cruzada, 'REF-001')
		self.assertEqual(producto.descripcion, 'Producto A')


class WarehouseStockTests(TestCase):
	def test_product_can_have_only_one_stock_record_per_warehouse(self):
		warehouse = Warehouse.objects.create(name='Almacen A', code='A')
		product = Producto.objects.create(nombre='WH-001', stock=10)
		ProductStock.objects.create(product=product, warehouse=warehouse, quantity=10)

		with self.assertRaises(IntegrityError):
			with transaction.atomic():
				ProductStock.objects.create(product=product, warehouse=warehouse, quantity=1)

	def test_warehouse_stock_changes_keep_legacy_total_in_sync(self):
		warehouse = Warehouse.objects.create(name='Almacen B', code='B')
		product = Producto.objects.create(nombre='WH-002', stock=0)

		apply_warehouse_stock_change(product, warehouse, 8)
		product.refresh_from_db()
		stock = ProductStock.objects.get(product=product, warehouse=warehouse)
		self.assertEqual(stock.quantity, 8)
		self.assertEqual(product.stock, 8)

		with self.assertRaises(InsufficientWarehouseStock):
			apply_warehouse_stock_change(product, warehouse, -9)

		product.refresh_from_db()
		stock.refresh_from_db()
		self.assertEqual(stock.quantity, 8)
		self.assertEqual(product.stock, 8)

	def test_transfer_moves_stock_between_warehouses_without_changing_total(self):
		origin = Warehouse.objects.create(name='Almacen origen', code='ORIGEN')
		destination = Warehouse.objects.create(name='Almacen destino', code='DESTINO')
		product = Producto.objects.create(nombre='WH-TRANSFER', stock=0, cost=12)
		apply_warehouse_stock_change(product, origin, 8)

		transfer = create_warehouse_transfer(
			origin_warehouse=origin,
			destination_warehouse=destination,
			items=[{'product': product, 'quantity': 3, 'observation': 'Reposicion'}],
			user=None,
			description='Reposicion semanal',
		)

		product.refresh_from_db()
		self.assertEqual(product.stock, 8)
		self.assertEqual(ProductStock.objects.get(product=product, warehouse=origin).quantity, 5)
		self.assertEqual(ProductStock.objects.get(product=product, warehouse=destination).quantity, 3)
		self.assertEqual(transfer.movements.count(), 2)
		self.assertEqual(transfer.items.count(), 1)

	def test_transfer_with_insufficient_origin_stock_is_rolled_back(self):
		origin = Warehouse.objects.create(name='Almacen bajo', code='BAJO')
		destination = Warehouse.objects.create(name='Almacen alto', code='ALTO')
		product = Producto.objects.create(nombre='WH-TRANSFER-FAIL', stock=0)
		apply_warehouse_stock_change(product, origin, 2)

		with self.assertRaises(InsufficientWarehouseStock):
			create_warehouse_transfer(
				origin_warehouse=origin,
				destination_warehouse=destination,
				items=[{'product': product, 'quantity': 3}],
				user=None,
			)

		product.refresh_from_db()
		self.assertEqual(product.stock, 2)
		self.assertEqual(ProductStock.objects.get(product=product, warehouse=origin).quantity, 2)
		self.assertFalse(ProductStock.objects.filter(product=product, warehouse=destination).exists())

	def test_inventory_report_uses_selected_warehouse_stock(self):
		origin = Warehouse.objects.create(name='Almacen reporte A', code='REPORTE-A')
		destination = Warehouse.objects.create(name='Almacen reporte B', code='REPORTE-B')
		product = Producto.objects.create(nombre='WH-REPORT', stock=0)
		apply_warehouse_stock_change(product, origin, 2)
		apply_warehouse_stock_change(product, destination, 5)
		user = User.objects.create_user(
			username='warehouse-reporter',
			email='reporter@example.com',
			name='Warehouse Reporter',
			password='secret123',
			role=User.Roles.ALMACEN,
			default_warehouse=origin,
		)
		self.client.force_login(user)

		response = self.client.get(reverse('reporte_inventario'), {'warehouse_id': origin.id})

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context['selected_warehouse'], origin)
		self.assertEqual(response.context['total_productos'], 1)
		self.assertEqual(dict(response.context['brand_summary'])['Sin Marca']['stock'], 2)

	def test_pre_inventory_uses_selected_warehouse_stock(self):
		origin = Warehouse.objects.create(name='Almacen preinventario A', code='PRE-A')
		destination = Warehouse.objects.create(name='Almacen preinventario B', code='PRE-B')
		product = Producto.objects.create(nombre='WH-PREINVENTORY', stock=0)
		apply_warehouse_stock_change(product, origin, 3, location='A-1')
		apply_warehouse_stock_change(product, destination, 6, location='B-1')
		user = User.objects.create_user(
			username='preinventory-reporter',
			email='preinventory@example.com',
			name='Preinventory Reporter',
			password='secret123',
			role=User.Roles.ALMACEN,
			default_warehouse=origin,
		)
		self.client.force_login(user)

		response = self.client.get(reverse('pre_inventario'), {'warehouse_id': origin.id, 'con_stock': 'on'})

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context['selected_warehouse'], origin)
		self.assertEqual(response.context['productos'].object_list[0].report_stock, 3)
		self.assertEqual(response.context['productos'].object_list[0].report_location, 'A-1')

	def test_analytics_uses_selected_warehouse_stock(self):
		warehouse = Warehouse.objects.create(name='Almacen analitica', code='ANALITICA')
		user = User.objects.create_user(
			username='analytics-reporter',
			email='analytics@example.com',
			name='Analytics Reporter',
			password='secret123',
			role=User.Roles.ALMACEN,
			default_warehouse=warehouse,
		)
		self.client.force_login(user)

		response = self.client.get(reverse('reporte_analitica_productos'), {'warehouse_id': warehouse.id})

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context['selected_warehouse'], warehouse)

	def test_warehouse_user_cannot_resolve_another_warehouse(self):
		assigned = Warehouse.objects.create(name='Almacen asignado', code='ASIGNADO')
		other = Warehouse.objects.create(name='Almacen ajeno', code='AJENO')
		user = User.objects.create_user(
			username='warehouse-user',
			email='warehouse-user@example.com',
			name='Warehouse User',
			password='secret123',
			role=User.Roles.ALMACEN,
			default_warehouse=assigned,
		)

		self.assertEqual(resolve_user_warehouse(user), assigned)
		with self.assertRaises(WarehouseAccessDenied):
			resolve_user_warehouse(user, other.id)
